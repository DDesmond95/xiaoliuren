"""
xiao_liuren_excel.py

A comprehensive Xiao Liu Ren (小六壬) calculator and Excel workbook generator.

This module is designed for two audiences at once:

1. Users who want a correct and structured implementation of standard
   Xiao Liu Ren calculation logic.
2. Beginners who know nothing about Xiao Liu Ren and need the exported
   Excel workbook itself to explain what is happening.

Features
--------
- Gregorian datetime -> lunar month/day conversion
- Civil hour -> Earthly Branch (地支时) conversion
- Standard Xiao Liu Ren calculation
- Optional late-Zi-hour rollover handling
- Optional leap-lunar-month handling policy
- Dynamic beginner-friendly explanations for each result row
- Excel workbook export with multiple educational sheets
- Summary statistics and glossary

Important scope note
--------------------
This module implements the compact "Xiao Liu Ren" method only.

It does NOT implement:
- Da Liu Ren (大六壬)
- Qi Men Dun Jia (奇门遁甲)
- other Chinese divination systems

Core rule implemented
---------------------
The standard compact Xiao Liu Ren rule is often described as:

    大安起正月，月上起日，日上起时

Operationally, this means:
1. Start from 大安 for lunar month 1
2. Count by lunar month
3. From that position count by lunar day
4. From that position count by Earthly Branch hour
5. The final palace is the result

The six results are, in order:

    1. 大安
    2. 留连
    3. 速喜
    4. 赤口
    5. 小吉
    6. 空亡

Equivalent compact formula:

    index = (lunar_month + lunar_day + earthly_hour_number - 3) % 6

where:
- lunar_month is 1-based
- lunar_day is 1-based
- earthly_hour_number is 1-based with 子=1, 丑=2, ..., 亥=12

Dependencies
------------
- lunardate
- openpyxl

Install:
    pip install lunardate openpyxl

Pylint design notes
-------------------
This module uses:
- explicit constants
- typed dataclasses
- docstrings
- small functions with limited responsibilities
- named policies via Enum
- no wildcard imports
"""

# pylint: disable=too-many-lines

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from enum import Enum
from functools import lru_cache
from typing import Dict, List, Sequence, Tuple

from lunardate import LunarDate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# =============================================================================
# Configuration enums
# =============================================================================


class LeapMonthPolicy(str, Enum):
    """
    Policy for handling leap lunar months (闰月).

    SAME_MONTH:
        Treat the leap month as its nominal month number.
        Example: leap 4th month -> 4

    NEXT_MONTH:
        Treat the leap month as the next month number.
        Example: leap 4th month -> 5
        If leap 12th month occurs, wrap to 1.

    ERROR:
        Raise an exception if a leap month is encountered.

    Notes
    -----
    Different lineages may handle leap months differently. There is no single
    guaranteed universal convention across all modern practitioners.
    """

    SAME_MONTH = "same_month"
    NEXT_MONTH = "next_month"
    ERROR = "error"


class ZiHourRolloverPolicy(str, Enum):
    """
    Policy for handling late Zi hour (23:00-23:59).

    NONE:
        Use the Gregorian date as-is.

    LATE_ZI_TO_NEXT_DAY:
        Treat 23:00-23:59 as belonging to the next day for lunar date
        conversion, while the hour still remains 子时 for branch purposes.

    Notes
    -----
    Some traditions distinguish early and late 子时. This module implements
    a common practical variant only.
    """

    NONE = "none"
    LATE_ZI_TO_NEXT_DAY = "late_zi_to_next_day"


# =============================================================================
# Constants and domain data
# =============================================================================

XIAO_LIUREN_RESULTS: Tuple[str, ...] = (
    "大安",
    "留连",
    "速喜",
    "赤口",
    "小吉",
    "空亡",
)

EARTHLY_BRANCHES: Tuple[str, ...] = (
    "子",
    "丑",
    "寅",
    "卯",
    "辰",
    "巳",
    "午",
    "未",
    "申",
    "酉",
    "戌",
    "亥",
)

EARTHLY_BRANCH_HOUR_RANGES: Dict[str, str] = {
    "子": "23:00-00:59",
    "丑": "01:00-02:59",
    "寅": "03:00-04:59",
    "卯": "05:00-06:59",
    "辰": "07:00-08:59",
    "巳": "09:00-10:59",
    "午": "11:00-12:59",
    "未": "13:00-14:59",
    "申": "15:00-16:59",
    "酉": "17:00-18:59",
    "戌": "19:00-20:59",
    "亥": "21:00-22:59",
}

RESULT_DETAILS: Dict[str, Dict[str, str]] = {
    "大安": {
        "basic_nature": "Stable, calm, safe, settled, broadly favorable.",
        "movement": "Stillness or low movement.",
        "tone": "Generally auspicious.",
        "plain_explanation": (
            "This usually suggests stability, safety, and a steady outcome. "
            "It tends to favor patience, calm handling, and ordinary progress."
        ),
        "beginner_tip": (
            "If you know nothing about Xiao Liu Ren, think of 大安 as "
            "'things are relatively stable and not chaotic'."
        ),
        "advice": (
            "Proceed steadily. Good for ordinary matters, staying put, "
            "planning, and preserving what is already working."
        ),
    },
    "留连": {
        "basic_nature": "Delay, repetition, entanglement, slowness.",
        "movement": "Slow movement; matter lingers.",
        "tone": "Mixed; usually obstructed rather than outright disastrous.",
        "plain_explanation": (
            "This usually suggests delay, repetition, back-and-forth, or "
            "difficulty getting a clean result quickly."
        ),
        "beginner_tip": (
            "Treat 留连 as 'it gets dragged out' rather than a clean yes/no."
        ),
        "advice": (
            "Expect waiting, follow-up, and possible rework. Avoid assuming "
            "the first attempt will fully solve the issue."
        ),
    },
    "速喜": {
        "basic_nature": "Fast movement, quick response, happy news.",
        "movement": "Fast movement.",
        "tone": "Generally auspicious.",
        "plain_explanation": (
            "This usually suggests quick developments, messages, responses, "
            "or an encouraging turn."
        ),
        "beginner_tip": ("Treat 速喜 as 'something favorable may happen quickly'."),
        "advice": (
            "Act promptly, watch for incoming news, and be ready to respond "
            "because timing may matter."
        ),
    },
    "赤口": {
        "basic_nature": "Conflict, harshness, injury risk, arguments.",
        "movement": "Movement with friction or trouble.",
        "tone": "Generally unfavorable.",
        "plain_explanation": (
            "This usually suggests disagreement, tension, verbal conflict, "
            "or some sharper and less harmonious influence."
        ),
        "beginner_tip": (
            "Treat 赤口 as 'watch out for conflict, arguments, or roughness'."
        ),
        "advice": (
            "Avoid escalation, be precise in communication, and reduce risk "
            "of disputes, mistakes, or careless action."
        ),
    },
    "小吉": {
        "basic_nature": "Small good fortune, modest favorability, useful help.",
        "movement": "Gentle or moderate movement.",
        "tone": "Auspicious, though usually softer than 大安.",
        "plain_explanation": (
            "This usually suggests a reasonably good outcome, helpful support, "
            "or a manageable and favorable direction."
        ),
        "beginner_tip": ("Treat 小吉 as 'good enough, modestly favorable, workable'."),
        "advice": (
            "Proceed with moderate confidence. This often supports practical "
            "progress, cooperation, or gradual benefit."
        ),
    },
    "空亡": {
        "basic_nature": "Emptiness, non-arrival, weakness, unreliability.",
        "movement": "Unstable or hollow movement.",
        "tone": "Usually unfavorable or inconclusive.",
        "plain_explanation": (
            "This usually suggests emptiness, lack of substance, failure to "
            "materialize, or unreliable appearances."
        ),
        "beginner_tip": ("Treat 空亡 as 'something may not really hold together'."),
        "advice": (
            "Verify facts, avoid overconfidence, and do not assume the matter "
            "will fully materialize just because it looks promising."
        ),
    },
}

COMMON_MEANING_TAGS: Dict[str, Tuple[str, str, str, str, str]] = {
    "大安": ("静", "不动", "吉利", "阳", "木"),
    "留连": ("慢", "反复", "拖延", "阴", "土"),
    "速喜": ("动", "快速", "吉利", "阳", "火"),
    "赤口": ("吵", "意见不和", "见金属利器伤害类", "阴", "金"),
    "小吉": ("动", "缓慢", "吉利", "阳", "水"),
    "空亡": ("无", "跟随性", "遇好则好，遇坏则坏", "阴", "土"),
}

RESULT_FILL_COLORS: Dict[str, str] = {
    "大安": "D9EAD3",
    "留连": "FFF2CC",
    "速喜": "D9EAF7",
    "赤口": "F4CCCC",
    "小吉": "EAD1DC",
    "空亡": "E6E6E6",
}


# =============================================================================
# Data models
# =============================================================================


@dataclass(frozen=True)
class LunarMonthDay:
    """
    Lunar month/day payload used in Xiao Liu Ren calculation.

    Attributes
    ----------
    lunar_year:
        Lunar year returned by the conversion library.

    lunar_month:
        Effective lunar month after leap-month policy has been applied.

    lunar_day:
        Lunar day, 1-based.

    is_leap_month:
        Whether the original lunar month was a leap month.

    raw_lunar_month:
        The original lunar month number returned by the library before any
        policy transformation.
    """

    lunar_year: int
    lunar_month: int
    lunar_day: int
    is_leap_month: bool
    raw_lunar_month: int


@dataclass(frozen=True)
class EarthlyBranchHour:
    """
    Earthly Branch hour representation.

    Attributes
    ----------
    branch_name:
        The branch symbol, such as 子 or 辰.

    branch_number:
        The 1-based number used in the Xiao Liu Ren calculation.
        子=1, 丑=2, ..., 亥=12

    branch_range:
        Human-readable modern clock range for the branch.
    """

    branch_name: str
    branch_number: int
    branch_range: str


@dataclass(frozen=True)
class XiaoLiuRenReading:
    """
    A single fully resolved Xiao Liu Ren reading.

    Attributes
    ----------
    original_datetime:
        The original Gregorian datetime supplied by the user.

    adjusted_datetime:
        Datetime after optional late-Zi-hour rollover policy.

    lunar_year:
        Lunar year used.

    lunar_month:
        Effective lunar month used in the final calculation.

    lunar_day:
        Lunar day used.

    is_leap_month:
        Whether the original lunar month was leap.

    raw_lunar_month:
        Original lunar month before policy handling.

    earthly_branch:
        Earthly Branch name for the hour.

    earthly_branch_number:
        1-based Earthly Branch number.

    earthly_branch_range:
        Human-readable modern hour range for the branch.

    formula_index:
        The zero-based modular index used internally.

    result:
        Final result among the six Xiao Liu Ren outcomes.

    meaning_1..meaning_5:
        Common simplified meaning tags.

    beginner_summary:
        Plain-language summary for someone unfamiliar with the system.

    detailed_explanation:
        Dynamic explanation describing how the result was calculated and what
        it means in beginner-friendly language.

    advice:
        Practical reading tip tied to the result.

    calculation_steps:
        Compact textual breakdown of the formula.
    """

    original_datetime: datetime
    adjusted_datetime: datetime
    lunar_year: int
    lunar_month: int
    lunar_day: int
    is_leap_month: bool
    raw_lunar_month: int
    earthly_branch: str
    earthly_branch_number: int
    earthly_branch_range: str
    formula_index: int
    result: str
    meaning_1: str
    meaning_2: str
    meaning_3: str
    meaning_4: str
    meaning_5: str
    beginner_summary: str
    detailed_explanation: str
    advice: str
    calculation_steps: str


# =============================================================================
# Validation helpers
# =============================================================================


def validate_hour(hour_value: int) -> None:
    """
    Validate a 24-hour clock hour.

    Parameters
    ----------
    hour_value:
        Integer hour expected to be in 0..23.

    Raises
    ------
    ValueError
        If the hour is outside the allowed range.
    """
    if not 0 <= hour_value <= 23:
        raise ValueError(f"Hour must be in range 0..23, got {hour_value!r}.")


def validate_datetime_range(
    start_datetime: datetime,
    end_datetime: datetime,
) -> None:
    """
    Validate that a datetime range is correctly ordered.

    Parameters
    ----------
    start_datetime:
        Inclusive start datetime.

    end_datetime:
        Inclusive end datetime.

    Raises
    ------
    ValueError
        If start_datetime is after end_datetime.
    """
    if start_datetime > end_datetime:
        raise ValueError("start_datetime must be less than or equal to end_datetime.")


# =============================================================================
# Core date and lunar conversion logic
# =============================================================================


def apply_zi_hour_rollover(
    current_datetime: datetime,
    policy: ZiHourRolloverPolicy,
) -> datetime:
    """
    Apply optional late-Zi-hour date rollover.

    Parameters
    ----------
    current_datetime:
        Original datetime.

    policy:
        Zi-hour rollover policy.

    Returns
    -------
    datetime
        Possibly adjusted datetime for lunar-date conversion.
    """
    validate_hour(current_datetime.hour)

    if (
        policy == ZiHourRolloverPolicy.LATE_ZI_TO_NEXT_DAY
        and current_datetime.hour == 23
    ):
        return current_datetime + timedelta(days=1)

    return current_datetime


def resolve_leap_month(
    lunar_month: int,
    is_leap_month: bool,
    policy: LeapMonthPolicy,
) -> int:
    """
    Resolve the effective lunar month according to the selected policy.

    Parameters
    ----------
    lunar_month:
        Raw lunar month number returned by the conversion library.

    is_leap_month:
        Whether the source month is leap.

    policy:
        Leap-month handling policy.

    Returns
    -------
    int
        Effective lunar month number used in the divination calculation.

    Raises
    ------
    ValueError
        If the policy is ERROR and a leap month is encountered.
    """
    if not is_leap_month:
        return lunar_month

    if policy == LeapMonthPolicy.SAME_MONTH:
        return lunar_month

    if policy == LeapMonthPolicy.NEXT_MONTH:
        return 1 if lunar_month == 12 else lunar_month + 1

    if policy == LeapMonthPolicy.ERROR:
        raise ValueError(
            f"Encountered leap lunar month {lunar_month}, "
            "and leap month policy is ERROR."
        )

    raise ValueError(f"Unsupported leap month policy: {policy!r}.")


@lru_cache(maxsize=10000)
def solar_to_lunar_month_day_cached(
    year_value: int,
    month_value: int,
    day_value: int,
    leap_month_policy_value: str,
) -> LunarMonthDay:
    """
    Cached Gregorian-date to lunar-month/day conversion.

    Parameters
    ----------
    year_value:
        Gregorian year.

    month_value:
        Gregorian month.

    day_value:
        Gregorian day.

    leap_month_policy_value:
        String value of the LeapMonthPolicy enum.

    Returns
    -------
    LunarMonthDay
        Structured lunar date payload.

    Notes
    -----
    The cache is useful because large hourly ranges contain many repeated
    dates, and lunar conversion is only date-based in this implementation.
    """
    lunar_date = LunarDate.fromSolarDate(year_value, month_value, day_value)

    raw_lunar_month = int(lunar_date.month)
    lunar_day = int(lunar_date.day)
    lunar_year = int(lunar_date.year)

    is_leap_month = bool(getattr(lunar_date, "isLeapMonth", False))

    effective_month = resolve_leap_month(
        lunar_month=raw_lunar_month,
        is_leap_month=is_leap_month,
        policy=LeapMonthPolicy(leap_month_policy_value),
    )

    return LunarMonthDay(
        lunar_year=lunar_year,
        lunar_month=effective_month,
        lunar_day=lunar_day,
        is_leap_month=is_leap_month,
        raw_lunar_month=raw_lunar_month,
    )


def solar_to_lunar_month_day(
    current_date: date,
    leap_month_policy: LeapMonthPolicy = LeapMonthPolicy.SAME_MONTH,
) -> LunarMonthDay:
    """
    Convert a Gregorian date to the lunar month/day used by Xiao Liu Ren.

    Parameters
    ----------
    current_date:
        Gregorian date to convert.

    leap_month_policy:
        Policy for handling leap lunar months.

    Returns
    -------
    LunarMonthDay
        Structured lunar date result.
    """
    return solar_to_lunar_month_day_cached(
        current_date.year,
        current_date.month,
        current_date.day,
        leap_month_policy.value,
    )


def get_earthly_branch_hour(hour_value: int) -> EarthlyBranchHour:
    """
    Convert a civil hour to Earthly Branch hour.

    Parameters
    ----------
    hour_value:
        Civil hour in 0..23.

    Returns
    -------
    EarthlyBranchHour
        Branch symbol, 1-based number, and readable time range.

    Mapping
    -------
    23:00-00:59 -> 子 -> 1
    01:00-02:59 -> 丑 -> 2
    03:00-04:59 -> 寅 -> 3
    ...
    21:00-22:59 -> 亥 -> 12
    """
    validate_hour(hour_value)

    index = ((hour_value + 1) // 2) % 12
    branch_name = EARTHLY_BRANCHES[index]

    return EarthlyBranchHour(
        branch_name=branch_name,
        branch_number=index + 1,
        branch_range=EARTHLY_BRANCH_HOUR_RANGES[branch_name],
    )


# =============================================================================
# Xiao Liu Ren calculation and explanation generation
# =============================================================================


def calculate_xiao_liuren_index(
    lunar_month: int,
    lunar_day: int,
    earthly_time_number: int,
) -> int:
    """
    Calculate the zero-based modular index for Xiao Liu Ren.

    Parameters
    ----------
    lunar_month:
        Effective lunar month, 1-based.

    lunar_day:
        Lunar day, 1-based.

    earthly_time_number:
        Earthly Branch hour number, 1-based.

    Returns
    -------
    int
        Zero-based index into XIAO_LIUREN_RESULTS.
    """
    if lunar_month < 1:
        raise ValueError(f"lunar_month must be >= 1, got {lunar_month!r}.")
    if lunar_day < 1:
        raise ValueError(f"lunar_day must be >= 1, got {lunar_day!r}.")
    if earthly_time_number < 1:
        raise ValueError(
            "earthly_time_number must be >= 1, " f"got {earthly_time_number!r}."
        )

    return (lunar_month + lunar_day + earthly_time_number - 3) % 6


def calculate_xiao_liuren_result(
    lunar_month: int,
    lunar_day: int,
    earthly_time_number: int,
) -> str:
    """
    Calculate the final Xiao Liu Ren result.

    Parameters
    ----------
    lunar_month:
        Effective lunar month, 1-based.

    lunar_day:
        Lunar day, 1-based.

    earthly_time_number:
        Earthly Branch hour number, 1-based.

    Returns
    -------
    str
        One of the six Xiao Liu Ren result names.
    """
    result_index = calculate_xiao_liuren_index(
        lunar_month=lunar_month,
        lunar_day=lunar_day,
        earthly_time_number=earthly_time_number,
    )
    return XIAO_LIUREN_RESULTS[result_index]


def get_common_meaning_tags(result_name: str) -> Tuple[str, str, str, str, str]:
    """
    Return simplified common tags for a result.

    Parameters
    ----------
    result_name:
        One of the six Xiao Liu Ren results.

    Returns
    -------
    tuple[str, str, str, str, str]
        Meaning tags. Returns empty strings if not found.
    """
    return COMMON_MEANING_TAGS.get(result_name, ("", "", "", "", ""))


def build_beginner_summary(
    result_name: str,
    lunar_month: int,
    lunar_day: int,
    earthly_branch_name: str,
) -> str:
    """
    Build a plain-language beginner summary for a single reading.

    Parameters
    ----------
    result_name:
        Final Xiao Liu Ren result.

    lunar_month:
        Effective lunar month used.

    lunar_day:
        Lunar day used.

    earthly_branch_name:
        Earthly Branch symbol for the hour.

    Returns
    -------
    str
        Beginner-friendly one-paragraph summary.
    """
    details = RESULT_DETAILS[result_name]
    return (
        f"For lunar month {lunar_month}, lunar day {lunar_day}, and "
        f"{earthly_branch_name} hour, the result is {result_name}. "
        f"In simple terms, this result suggests: "
        f"{details['plain_explanation']} "
        f"{details['beginner_tip']}"
    )


def build_calculation_steps(
    lunar_month: int,
    lunar_day: int,
    earthly_branch_number: int,
    result_index: int,
    result_name: str,
) -> str:
    """
    Build a compact textual formula breakdown for the reading.

    Parameters
    ----------
    lunar_month:
        Effective lunar month used.

    lunar_day:
        Lunar day used.

    earthly_branch_number:
        1-based Earthly Branch number used.

    result_index:
        Zero-based modular result index.

    result_name:
        Final result name.

    Returns
    -------
    str
        Formula explanation string.
    """
    raw_total = lunar_month + lunar_day + earthly_branch_number - 3
    return (
        f"Formula: (month + day + hour - 3) % 6 = "
        f"({lunar_month} + {lunar_day} + {earthly_branch_number} - 3) % 6 = "
        f"{raw_total} % 6 = {result_index}. "
        f"Index {result_index} maps to {result_name}."
    )


def build_detailed_explanation(
    result_name: str,
    lunar_month: int,
    lunar_day: int,
    earthly_branch_name: str,
    earthly_branch_range: str,
    earthly_branch_number: int,
    result_index: int,
) -> str:
    """
    Build a dynamic detailed explanation suitable for Excel output.

    Parameters
    ----------
    result_name:
        Final result.

    lunar_month:
        Effective lunar month.

    lunar_day:
        Lunar day.

    earthly_branch_name:
        Earthly Branch symbol.

    earthly_branch_range:
        Human-readable clock range.

    earthly_branch_number:
        1-based branch number.

    result_index:
        Zero-based modular index.

    Returns
    -------
    str
        A dynamic, beginner-friendly explanation paragraph.
    """
    details = RESULT_DETAILS[result_name]

    return (
        f"This reading uses the standard Xiao Liu Ren rule: start from 大安 "
        f"for lunar month 1, then count by lunar month, then by lunar day, "
        f"then by Earthly Branch hour. Here the effective lunar month is "
        f"{lunar_month}, the lunar day is {lunar_day}, and the hour is "
        f"{earthly_branch_name} hour ({earthly_branch_range}), which is "
        f"hour number {earthly_branch_number}. Using the formula "
        f"(month + day + hour - 3) % 6 gives index {result_index}, which maps "
        f"to {result_name}. In interpretation, {result_name} is commonly "
        f"understood as: {details['basic_nature']} The movement quality is: "
        f"{details['movement']} The overall tone is: {details['tone']} "
        f"For a complete beginner, a simple reading is: "
        f"{details['plain_explanation']}"
    )


def build_xiao_liuren_reading(
    current_datetime: datetime,
    leap_month_policy: LeapMonthPolicy = LeapMonthPolicy.SAME_MONTH,
    zi_hour_rollover_policy: ZiHourRolloverPolicy = ZiHourRolloverPolicy.NONE,
) -> XiaoLiuRenReading:
    """
    Build a full Xiao Liu Ren reading from a Gregorian datetime.

    Parameters
    ----------
    current_datetime:
        Datetime to evaluate.

    leap_month_policy:
        Leap-month handling policy.

    zi_hour_rollover_policy:
        Late-Zi-hour date rollover policy.

    Returns
    -------
    XiaoLiuRenReading
        Fully resolved reading object.
    """
    adjusted_datetime = apply_zi_hour_rollover(
        current_datetime=current_datetime,
        policy=zi_hour_rollover_policy,
    )

    lunar_info = solar_to_lunar_month_day(
        current_date=adjusted_datetime.date(),
        leap_month_policy=leap_month_policy,
    )

    branch_info = get_earthly_branch_hour(current_datetime.hour)

    result_index = calculate_xiao_liuren_index(
        lunar_month=lunar_info.lunar_month,
        lunar_day=lunar_info.lunar_day,
        earthly_time_number=branch_info.branch_number,
    )
    result_name = XIAO_LIUREN_RESULTS[result_index]

    meaning_1, meaning_2, meaning_3, meaning_4, meaning_5 = get_common_meaning_tags(
        result_name
    )

    beginner_summary = build_beginner_summary(
        result_name=result_name,
        lunar_month=lunar_info.lunar_month,
        lunar_day=lunar_info.lunar_day,
        earthly_branch_name=branch_info.branch_name,
    )

    detailed_explanation = build_detailed_explanation(
        result_name=result_name,
        lunar_month=lunar_info.lunar_month,
        lunar_day=lunar_info.lunar_day,
        earthly_branch_name=branch_info.branch_name,
        earthly_branch_range=branch_info.branch_range,
        earthly_branch_number=branch_info.branch_number,
        result_index=result_index,
    )

    advice = RESULT_DETAILS[result_name]["advice"]

    calculation_steps = build_calculation_steps(
        lunar_month=lunar_info.lunar_month,
        lunar_day=lunar_info.lunar_day,
        earthly_branch_number=branch_info.branch_number,
        result_index=result_index,
        result_name=result_name,
    )

    return XiaoLiuRenReading(
        original_datetime=current_datetime,
        adjusted_datetime=adjusted_datetime,
        lunar_year=lunar_info.lunar_year,
        lunar_month=lunar_info.lunar_month,
        lunar_day=lunar_info.lunar_day,
        is_leap_month=lunar_info.is_leap_month,
        raw_lunar_month=lunar_info.raw_lunar_month,
        earthly_branch=branch_info.branch_name,
        earthly_branch_number=branch_info.branch_number,
        earthly_branch_range=branch_info.branch_range,
        formula_index=result_index,
        result=result_name,
        meaning_1=meaning_1,
        meaning_2=meaning_2,
        meaning_3=meaning_3,
        meaning_4=meaning_4,
        meaning_5=meaning_5,
        beginner_summary=beginner_summary,
        detailed_explanation=detailed_explanation,
        advice=advice,
        calculation_steps=calculation_steps,
    )


# =============================================================================
# Range generation
# =============================================================================


def generate_hourly_datetimes(
    start_datetime: datetime,
    end_datetime: datetime,
) -> List[datetime]:
    """
    Generate an inclusive list of hourly datetimes.

    Parameters
    ----------
    start_datetime:
        Inclusive start datetime.

    end_datetime:
        Inclusive end datetime.

    Returns
    -------
    list[datetime]
        Hourly datetimes.
    """
    validate_datetime_range(start_datetime, end_datetime)

    datetimes: List[datetime] = []
    current_datetime = start_datetime

    while current_datetime <= end_datetime:
        datetimes.append(current_datetime)
        current_datetime += timedelta(hours=1)

    return datetimes


def generate_xiao_liuren_readings(
    start_datetime: datetime,
    end_datetime: datetime,
    leap_month_policy: LeapMonthPolicy = LeapMonthPolicy.SAME_MONTH,
    zi_hour_rollover_policy: ZiHourRolloverPolicy = ZiHourRolloverPolicy.NONE,
) -> List[XiaoLiuRenReading]:
    """
    Generate Xiao Liu Ren readings for each hour in a datetime range.

    Parameters
    ----------
    start_datetime:
        Inclusive start datetime.

    end_datetime:
        Inclusive end datetime.

    leap_month_policy:
        Leap-month handling policy.

    zi_hour_rollover_policy:
        Late-Zi-hour rollover policy.

    Returns
    -------
    list[XiaoLiuRenReading]
        Generated readings.
    """
    readings: List[XiaoLiuRenReading] = []

    for current_datetime in generate_hourly_datetimes(
        start_datetime=start_datetime,
        end_datetime=end_datetime,
    ):
        reading = build_xiao_liuren_reading(
            current_datetime=current_datetime,
            leap_month_policy=leap_month_policy,
            zi_hour_rollover_policy=zi_hour_rollover_policy,
        )
        readings.append(reading)

    return readings


# =============================================================================
# Excel-writing helpers
# =============================================================================


def apply_header_style(worksheet, row_index: int = 1) -> None:
    """
    Apply a standard header style to a worksheet row.

    Parameters
    ----------
    worksheet:
        OpenPyXL worksheet object.

    row_index:
        Header row number.
    """
    fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in worksheet[row_index]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def autofit_column_widths(
    worksheet,
    min_width: int = 10,
    max_width: int = 60,
) -> None:
    """
    Auto-adjust column widths based on content length.

    Parameters
    ----------
    worksheet:
        OpenPyXL worksheet object.

    min_width:
        Minimum width to apply.

    max_width:
        Maximum width to apply.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                cell_value = "" if cell.value is None else str(cell.value)
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except TypeError:
                continue

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def freeze_top_row(worksheet) -> None:
    """
    Freeze the top row of a worksheet.

    Parameters
    ----------
    worksheet:
        OpenPyXL worksheet object.
    """
    worksheet.freeze_panes = "A2"


def apply_wrap_text_to_all_cells(worksheet) -> None:
    """
    Apply wrap-text alignment to all populated cells in a worksheet.

    Parameters
    ----------
    worksheet:
        OpenPyXL worksheet object.
    """
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def color_result_column(worksheet, result_column_index: int) -> None:
    """
    Color the result column cells based on the Xiao Liu Ren result value.

    Parameters
    ----------
    worksheet:
        OpenPyXL worksheet object.

    result_column_index:
        1-based column index containing the result name.
    """
    for row_index in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_index, column=result_column_index)
        result_name = str(cell.value)

        fill_color = RESULT_FILL_COLORS.get(result_name)
        if fill_color:
            cell.fill = PatternFill(
                fill_type="solid",
                start_color=fill_color,
                end_color=fill_color,
            )


# =============================================================================
# Excel sheet writers
# =============================================================================


def write_overview_sheet(
    workbook: Workbook,
    readings: Sequence[XiaoLiuRenReading],
) -> None:
    """
    Write the workbook Overview sheet.

    Parameters
    ----------
    workbook:
        OpenPyXL workbook.

    readings:
        Generated readings.
    """
    worksheet = workbook.active
    worksheet.title = "Overview"

    rows = [
        ["Item", "Value"],
        ["Workbook Purpose", "Xiao Liu Ren calculator and beginner-friendly explainer"],
        ["Method", "Standard compact Xiao Liu Ren only"],
        ["Formula", "(lunar_month + lunar_day + earthly_hour_number - 3) % 6"],
        ["Result Count", len(readings)],
        ["Six Results", "大安, 留连, 速喜, 赤口, 小吉, 空亡"],
        [
            "What this workbook includes",
            (
                "Overview, beginner instructions, settings, full results, "
                "summary counts, and glossary"
            ),
        ],
        [
            "Important caution",
            (
                "This workbook explains the standard compact method. "
                "Different schools may differ on leap months, late Zi hour, "
                "and interpretation details."
            ),
        ],
    ]

    for row in rows:
        worksheet.append(row)

    apply_header_style(worksheet)
    apply_wrap_text_to_all_cells(worksheet)
    autofit_column_widths(worksheet, max_width=80)


def write_how_to_read_sheet(workbook: Workbook) -> None:
    """
    Write a sheet explaining how a beginner should read the workbook.

    Parameters
    ----------
    workbook:
        OpenPyXL workbook.
    """
    worksheet = workbook.create_sheet(title="How_To_Read")

    rows = [
        ["Section", "Explanation"],
        [
            "What is Xiao Liu Ren?",
            (
                "Xiao Liu Ren is a compact Chinese divination method that "
                "reduces a time to one of six results: 大安, 留连, 速喜, "
                "赤口, 小吉, 空亡."
            ),
        ],
        [
            "What inputs are used?",
            (
                "This workbook uses lunar month, lunar day, and Earthly Branch "
                "hour. It does not use the Gregorian month/day directly for "
                "the final calculation."
            ),
        ],
        [
            "How is the result calculated?",
            (
                "The standard rule is: start from 大安 for lunar month 1, "
                "count by lunar month, then lunar day, then Earthly Branch "
                "hour. The compact formula is "
                "(month + day + hour - 3) % 6."
            ),
        ],
        [
            "What should a beginner look at first?",
            (
                "Start with the Results sheet. Read the columns Result, "
                "BeginnerSummary, DetailedExplanation, Advice, and "
                "CalculationSteps."
            ),
        ],
        [
            "What does a 'good' result mean?",
            (
                "Generally, 大安, 速喜, and 小吉 are read as more favorable. "
                "留连 often means delay. 赤口 often warns of conflict. "
                "空亡 often points to emptiness or unreliability."
            ),
        ],
        [
            "What are common caveats?",
            (
                "Interpretive traditions vary. Leap lunar month handling and "
                "late Zi-hour rollover are not always treated the same in "
                "every lineage."
            ),
        ],
        [
            "How should I use this workbook?",
            (
                "Use it first as a structured calculator and educational tool. "
                "Do not confuse it with Da Liu Ren, which is a much more "
                "complex and separate system."
            ),
        ],
    ]

    for row in rows:
        worksheet.append(row)

    apply_header_style(worksheet)
    apply_wrap_text_to_all_cells(worksheet)
    autofit_column_widths(worksheet, max_width=100)


def write_settings_sheet(
    workbook: Workbook,
    start_datetime: datetime,
    end_datetime: datetime,
    leap_month_policy: LeapMonthPolicy,
    zi_hour_rollover_policy: ZiHourRolloverPolicy,
) -> None:
    """
    Write a sheet documenting the workbook settings and assumptions.

    Parameters
    ----------
    workbook:
        OpenPyXL workbook.

    start_datetime:
        Range start.

    end_datetime:
        Range end.

    leap_month_policy:
        Selected leap-month policy.

    zi_hour_rollover_policy:
        Selected Zi-hour rollover policy.
    """
    worksheet = workbook.create_sheet(title="Settings")

    rows = [
        ["Setting", "Value"],
        ["StartDatetime", start_datetime.isoformat(sep=" ")],
        ["EndDatetime", end_datetime.isoformat(sep=" ")],
        ["LeapMonthPolicy", leap_month_policy.value],
        ["ZiHourRolloverPolicy", zi_hour_rollover_policy.value],
        [
            "Lunar source",
            "lunardate.LunarDate.fromSolarDate",
        ],
        [
            "Hour mapping",
            "23:00-00:59=子(1), 01:00-02:59=丑(2), ..., 21:00-22:59=亥(12)",
        ],
        [
            "Core formula",
            "(lunar_month + lunar_day + earthly_hour_number - 3) % 6",
        ],
        [
            "Assumption note",
            (
                "This workbook is designed around standard compact Xiao Liu Ren. "
                "Interpretation notes are educational summaries, not claims of "
                "universal doctrinal agreement."
            ),
        ],
    ]

    for row in rows:
        worksheet.append(row)

    apply_header_style(worksheet)
    apply_wrap_text_to_all_cells(worksheet)
    autofit_column_widths(worksheet, max_width=100)


def write_results_sheet(
    workbook: Workbook,
    readings: Sequence[XiaoLiuRenReading],
) -> None:
    """
    Write the detailed results sheet.

    Parameters
    ----------
    workbook:
        OpenPyXL workbook.

    readings:
        Generated readings.
    """
    worksheet = workbook.create_sheet(title="Results")

    headers = [
        "OriginalDatetime",
        "AdjustedDatetime",
        "GregorianYear",
        "GregorianMonth",
        "GregorianDay",
        "GregorianHour",
        "LunarYear",
        "LunarMonth",
        "LunarDay",
        "IsLeapMonth",
        "RawLunarMonth",
        "EarthlyBranch",
        "EarthlyBranchNumber",
        "EarthlyBranchRange",
        "FormulaIndex",
        "Result",
        "Meaning_1",
        "Meaning_2",
        "Meaning_3",
        "Meaning_4",
        "Meaning_5",
        "BeginnerSummary",
        "DetailedExplanation",
        "Advice",
        "CalculationSteps",
    ]
    worksheet.append(headers)

    for reading in readings:
        worksheet.append(
            [
                reading.original_datetime.isoformat(sep=" "),
                reading.adjusted_datetime.isoformat(sep=" "),
                reading.original_datetime.year,
                reading.original_datetime.month,
                reading.original_datetime.day,
                reading.original_datetime.hour,
                reading.lunar_year,
                reading.lunar_month,
                reading.lunar_day,
                reading.is_leap_month,
                reading.raw_lunar_month,
                reading.earthly_branch,
                reading.earthly_branch_number,
                reading.earthly_branch_range,
                reading.formula_index,
                reading.result,
                reading.meaning_1,
                reading.meaning_2,
                reading.meaning_3,
                reading.meaning_4,
                reading.meaning_5,
                reading.beginner_summary,
                reading.detailed_explanation,
                reading.advice,
                reading.calculation_steps,
            ]
        )

    apply_header_style(worksheet)
    freeze_top_row(worksheet)
    apply_wrap_text_to_all_cells(worksheet)
    autofit_column_widths(worksheet, max_width=80)

    # "Result" is the 16th column in the Results sheet.
    color_result_column(worksheet, result_column_index=16)


def write_summary_sheet(
    workbook: Workbook,
    readings: Sequence[XiaoLiuRenReading],
) -> None:
    """
    Write summary counts and percentages per result.

    Parameters
    ----------
    workbook:
        OpenPyXL workbook.

    readings:
        Generated readings.
    """
    worksheet = workbook.create_sheet(title="Summary")

    total_count = len(readings)
    counts: Dict[str, int] = {result_name: 0 for result_name in XIAO_LIUREN_RESULTS}

    for reading in readings:
        counts[reading.result] = counts.get(reading.result, 0) + 1

    worksheet.append(["Result", "Count", "Percentage", "BasicNature", "BeginnerTip"])

    for result_name in XIAO_LIUREN_RESULTS:
        count_value = counts[result_name]
        percentage = (count_value / total_count) if total_count else 0.0
        worksheet.append(
            [
                result_name,
                count_value,
                percentage,
                RESULT_DETAILS[result_name]["basic_nature"],
                RESULT_DETAILS[result_name]["beginner_tip"],
            ]
        )

    apply_header_style(worksheet)
    apply_wrap_text_to_all_cells(worksheet)
    autofit_column_widths(worksheet, max_width=80)
    color_result_column(worksheet, result_column_index=1)

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_index, column=3).number_format = "0.00%"


def write_glossary_sheet(workbook: Workbook) -> None:
    """
    Write a glossary sheet explaining all six results and the Earthly Branches.

    Parameters
    ----------
    workbook:
        OpenPyXL workbook.
    """
    worksheet = workbook.create_sheet(title="Glossary")

    worksheet.append(["Term", "Category", "Explanation", "Notes"])

    for result_name in XIAO_LIUREN_RESULTS:
        details = RESULT_DETAILS[result_name]
        worksheet.append(
            [
                result_name,
                "XiaoLiuRenResult",
                details["plain_explanation"],
                details["advice"],
            ]
        )

    for index_value, branch_name in enumerate(EARTHLY_BRANCHES, start=1):
        worksheet.append(
            [
                branch_name,
                "EarthlyBranchHour",
                (
                    f"{branch_name} hour corresponds to "
                    f"{EARTHLY_BRANCH_HOUR_RANGES[branch_name]}"
                ),
                f"Used as hour number {index_value} in the calculation.",
            ]
        )

    worksheet.append(
        [
            "Leap lunar month",
            "Concept",
            "A lunar month that is inserted in the lunisolar calendar.",
            "Different lineages may treat it differently in divination.",
        ]
    )
    worksheet.append(
        [
            "Late Zi hour",
            "Concept",
            "23:00-23:59, sometimes treated as the next day in some traditions.",
            "This workbook can optionally shift it for lunar-date conversion.",
        ]
    )
    worksheet.append(
        [
            "Formula index",
            "Concept",
            "The zero-based result of (month + day + hour - 3) % 6.",
            "It maps to the six Xiao Liu Ren results in fixed order.",
        ]
    )

    apply_header_style(worksheet)
    apply_wrap_text_to_all_cells(worksheet)
    autofit_column_widths(worksheet, max_width=90)


# =============================================================================
# Workbook orchestration
# =============================================================================


def export_xiao_liuren_to_excel(
    readings: Sequence[XiaoLiuRenReading],
    output_path: str,
    start_datetime: datetime,
    end_datetime: datetime,
    leap_month_policy: LeapMonthPolicy,
    zi_hour_rollover_policy: ZiHourRolloverPolicy,
) -> None:
    """
    Export Xiao Liu Ren readings to an educational Excel workbook.

    Parameters
    ----------
    readings:
        Generated readings.

    output_path:
        Target .xlsx file path.

    start_datetime:
        Range start.

    end_datetime:
        Range end.

    leap_month_policy:
        Leap-month policy used.

    zi_hour_rollover_policy:
        Zi-hour rollover policy used.
    """
    workbook = Workbook()

    write_overview_sheet(workbook, readings)
    write_how_to_read_sheet(workbook)
    write_settings_sheet(
        workbook=workbook,
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        leap_month_policy=leap_month_policy,
        zi_hour_rollover_policy=zi_hour_rollover_policy,
    )
    write_results_sheet(workbook, readings)
    write_summary_sheet(workbook, readings)
    write_glossary_sheet(workbook)

    workbook.save(output_path)


def export_readings_to_json(
    readings: Sequence[XiaoLiuRenReading],
    output_path: str,
) -> None:
    """
    Export Xiao Liu Ren readings to a JSON file for the web UI.

    This file is consumed by the GitHub Pages frontend so users can
    interactively query results without running Python.

    Parameters
    ----------
    readings:
        List of XiaoLiuRenReading objects generated by the calculator.

    output_path:
        Target JSON file path.
    """

    json_rows: List[Dict[str, object]] = []

    for reading in readings:

        json_rows.append(
            {
                # Full datetime string for display
                "datetime": reading.original_datetime.isoformat(sep=" "),
                # Hour used for quick matching in the UI
                "hour": reading.original_datetime.hour,
                # Lunar values used in the calculation
                "lunar_month": reading.lunar_month,
                "lunar_day": reading.lunar_day,
                # Earthly branch hour
                "branch": reading.earthly_branch,
                # Final Xiao Liu Ren result
                "result": reading.result,
                # Short meaning text
                "meaning": reading.beginner_summary,
                # Advice text for UI display
                "advice": reading.advice,
            }
        )

    with open(output_path, "w", encoding="utf-8") as file:
        json.dump(json_rows, file, ensure_ascii=False, indent=2)


# =============================================================================
# Testing helpers
# =============================================================================


def run_basic_self_tests() -> None:
    """
    Run a few basic sanity tests.

    These are not a full unit test suite, but they verify core logic:
    - Earthly Branch hour mapping
    - Formula behavior
    - A classical-style example:
      lunar 3rd month, lunar 5th day, 辰时(5) -> 小吉
    """
    assert get_earthly_branch_hour(23) == EarthlyBranchHour("子", 1, "23:00-00:59")
    assert get_earthly_branch_hour(0) == EarthlyBranchHour("子", 1, "23:00-00:59")
    assert get_earthly_branch_hour(1) == EarthlyBranchHour("丑", 2, "01:00-02:59")
    assert get_earthly_branch_hour(22) == EarthlyBranchHour("亥", 12, "21:00-22:59")

    assert calculate_xiao_liuren_result(1, 1, 1) == "大安"
    assert calculate_xiao_liuren_result(1, 1, 2) == "留连"
    assert calculate_xiao_liuren_result(3, 5, 5) == "小吉"


# =============================================================================
# Convenience entry point
# =============================================================================


def generate_and_export_xiao_liuren_workbook(
    start_datetime: datetime,
    end_datetime: datetime,
    output_path: str = "xiao_liuren_results.xlsx",
    json_output_path: str = "docs/xiao_liuren_data.json",
    leap_month_policy: LeapMonthPolicy = LeapMonthPolicy.SAME_MONTH,
    zi_hour_rollover_policy: ZiHourRolloverPolicy = ZiHourRolloverPolicy.NONE,
) -> List[XiaoLiuRenReading]:
    """
    Generate Xiao Liu Ren readings and export them to Excel.

    Parameters
    ----------
    start_datetime:
        Inclusive range start.

    end_datetime:
        Inclusive range end.

    output_path:
        Output Excel file path.

    leap_month_policy:
        Leap-month handling policy.

    zi_hour_rollover_policy:
        Late-Zi-hour rollover policy.

    Returns
    -------
    list[XiaoLiuRenReading]
        Generated readings.
    """
    readings = generate_xiao_liuren_readings(
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        leap_month_policy=leap_month_policy,
        zi_hour_rollover_policy=zi_hour_rollover_policy,
    )

    export_xiao_liuren_to_excel(
        readings=readings,
        output_path=output_path,
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        leap_month_policy=leap_month_policy,
        zi_hour_rollover_policy=zi_hour_rollover_policy,
    )

    export_readings_to_json(
        readings=readings,
        output_path=json_output_path,
    )

    return readings


def main() -> None:
    """
    Example main function.

    Adjust the date range and policies here as needed.
    """
    run_basic_self_tests()

    start_datetime = datetime(2026, 1, 1, 0, 0)
    end_datetime = datetime(2028, 1, 1, 0, 0)

    readings = generate_and_export_xiao_liuren_workbook(
        start_datetime=start_datetime,
        end_datetime=end_datetime,
        output_path="docs/xiao_liuren_results.xlsx",
        json_output_path="docs/xiao_liuren_data.json",
        leap_month_policy=LeapMonthPolicy.SAME_MONTH,
        zi_hour_rollover_policy=ZiHourRolloverPolicy.NONE,
    )

    print(
        f"Generated {len(readings)} readings and exported to "
        f"'xiao_liuren_results.xlsx'."
    )


if __name__ == "__main__":
    main()
